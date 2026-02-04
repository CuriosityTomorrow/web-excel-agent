from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from app.models.excel import Workbook as WorkbookModel
import io


class ExcelService:
    def __init__(self):
        self.workbooks = {}
        self.charts = {}  # Store charts by workbook_id

    def create_workbook(self, data: WorkbookModel) -> str:
        """Create a new workbook"""
        import uuid
        workbook_id = str(uuid.uuid4())
        data.id = workbook_id
        self.workbooks[workbook_id] = data
        return workbook_id

    def get_workbook(self, workbook_id: str) -> WorkbookModel:
        """Get workbook by ID"""
        return self.workbooks.get(workbook_id)

    def update_cell(self, workbook_id: str, sheet_index: int, row: int, col: int, value: str):
        """Update a cell value"""
        workbook = self.workbooks.get(workbook_id)
        if not workbook:
            raise ValueError(f"Workbook {workbook_id} not found")

        if sheet_index >= len(workbook.sheets):
            raise ValueError(f"Sheet index {sheet_index} out of range")

        sheet = workbook.sheets[sheet_index]

        # Ensure row exists
        while len(sheet.rows) <= row:
            sheet.rows.append([''] * len(sheet.rows[0]) if sheet.rows else [''])

        # Ensure column exists
        while len(sheet.rows[row]) <= col:
            sheet.rows[row].append('')

        sheet.rows[row][col] = value
        return workbook

    def export_to_excel(self, workbook_id: str) -> bytes:
        """Export workbook to Excel file"""
        workbook_data = self.workbooks.get(workbook_id)
        if not workbook_data:
            raise ValueError(f"Workbook {workbook_id} not found")

        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        for sheet_data in workbook_data.sheets:
            ws = wb.create_sheet(title=sheet_data.name)

            for row_idx, row in enumerate(sheet_data.rows, start=1):
                for col_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    # Add styling for header row
                    if row_idx == 1:
                        cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='left')

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    def add_sheet(self, workbook_id: str, sheet_name: str):
        """Add a new sheet to workbook"""
        workbook = self.workbooks.get(workbook_id)
        if not workbook:
            raise ValueError(f"Workbook {workbook_id} not found")

        from app.models.excel import Sheet
        new_sheet = Sheet(name=sheet_name, rows=[['', '', '', '']])
        workbook.sheets.append(new_sheet)
        return workbook

    def add_chart(self, workbook_id: str, chart: dict):
        """Add a chart to the workbook"""
        if workbook_id not in self.charts:
            self.charts[workbook_id] = []
        self.charts[workbook_id].append(chart)

    def set_workbook_charts(self, workbook_id: str, charts: list):
        """Replace all charts for a workbook (used during export)"""
        self.charts[workbook_id] = charts

    def sync_workbook(self, workbook_id: str, workbook_data: WorkbookModel):
        """Sync workbook data from frontend"""
        if workbook_id not in self.workbooks:
            raise ValueError(f"Workbook {workbook_id} not found")

        # Update the workbook data
        workbook_data.id = workbook_id
        self.workbooks[workbook_id] = workbook_data
        return workbook_data

    def export_to_excel_with_charts(self, workbook_id: str) -> bytes:
        """Export workbook to Excel file with charts"""
        workbook_data = self.workbooks.get(workbook_id)
        if not workbook_data:
            raise ValueError(f"Workbook {workbook_id} not found")

        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        for sheet_idx, sheet_data in enumerate(workbook_data.sheets):
            ws = wb.create_sheet(title=sheet_data.name)

            for row_idx, row in enumerate(sheet_data.rows, start=1):
                for col_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    # Add styling for header row
                    if row_idx == 1:
                        cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='left')

            # Add charts only to the first sheet
            if sheet_idx == 0 and workbook_id in self.charts:
                charts = self.charts[workbook_id]
                for idx, chart_data in enumerate(charts):
                    self._add_chart_to_sheet(ws, chart_data, len(sheet_data.rows), idx)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    def _add_chart_to_sheet(self, ws, chart_data, data_rows, chart_idx):
        """Add a chart to the worksheet"""
        chart_type = chart_data.get('type', 'bar')
        title = chart_data.get('title', 'Chart')
        data = chart_data.get('data', [])

        if not data or data_rows < 2:
            return

        # Create chart based on type
        if chart_type == 'bar':
            chart = BarChart()
            chart.type = 'col'
            chart.style = 10
            chart.y_axis.title = '数值'
            chart.x_axis.title = '类别'
        elif chart_type == 'line':
            chart = LineChart()
            chart.style = 10
            chart.y_axis.title = '数值'
            chart.x_axis.title = '类别'
        elif chart_type == 'pie':
            chart = PieChart()
            chart.style = 10
            # PieChart doesn't have axes
        else:
            chart = BarChart()
            chart.style = 10
            chart.y_axis.title = '数值'
            chart.x_axis.title = '类别'

        chart.title = title

        # Calculate data range - use column B for values (index 2)
        # and column A for categories (index 1)
        if data_rows > 1:
            # Find the column with numeric data (usually column B or C)
            value_col = 2  # B column
            cat_col = 1     # A column

            values_ref = Reference(ws, min_col=value_col, min_row=1, max_row=data_rows, max_col=value_col)
            cats_ref = Reference(ws, min_col=cat_col, min_row=2, max_row=data_rows)

            chart.add_data(values_ref, titles_from_data=True)
            chart.set_categories(cats_ref)

            # Position chart to the right of data
            # Position each chart vertically if there are multiple
            start_row = 2 + (chart_idx * 15)  # Offset for multiple charts
            chart.anchor = f'{chr(65 + value_col + 2)}{start_row}'  # Position 2 columns to the right
            chart.width = 15
            chart.height = 12

            ws.add_chart(chart)


# Global instance
excel_service = ExcelService()
